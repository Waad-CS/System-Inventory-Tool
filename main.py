import platform
import psutil
import pandas as pd
import datetime
import os
import socket
import subprocess
import math
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def get_detailed_model():
    try:
        # Fetching manufacturer and model accurately from the system
        manufacturer = subprocess.check_output('wmic computersystem get manufacturer').decode().split('\n', 1)[1].strip()
        model = subprocess.check_output('wmic computersystem get model').decode().split('\n', 1)[1].strip()
        return f"{manufacturer} {model}"
    except:
        return "Unknown Device"

def collect_inventory():
    print(f"--- Scanning Device: {platform.node()} ---")
    
    # RAM calculation and rounding up (e.g., 7.73 to 8)
    raw_ram = psutil.virtual_memory().total / (1024**3)
    formatted_ram = math.ceil(raw_ram)
    
    # Disk space calculation (C: drive)
    disk_info = psutil.disk_usage('C:')
    total_disk_gb = math.ceil(disk_info.total / (1024**3))
    
    try:
        ip_address = socket.gethostbyname(socket.gethostname())
    except:
        ip_address = "N/A"

    return {
        # Precise date and time for audit trail
        "Last Scan": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"), 
        "Status": "", 
        "Physical Location": "",
        "Device Name": platform.node(),
        "IP Address": ip_address,
        "Manufacturer & Model": get_detailed_model(),
        "Total RAM (GB)": formatted_ram,
        "Disk Space (GB)": total_disk_gb,
        "Notes": ""
    }

def save_report(new_data):
    filename = "Master_IT_Inventory.xlsx"
    
    if os.path.exists(filename):
        df_old = pd.read_excel(filename)
        device_name = str(new_data["Device Name"])
        
        if device_name in df_old["Device Name"].astype(str).values:
            # Updating an existing device
            idx = df_old.index[df_old["Device Name"].astype(str) == device_name][0]
            
            for column in new_data.keys():
                # Preserve manual entries like 'Physical Location' and 'Notes'
                if column not in ["Physical Location", "Notes"]:
                    df_old.at[idx, column] = new_data[column]
            
            df_old.at[idx, "Status"] = "Updated"
            df_final = df_old
            print(f"🔄 Device updated: {device_name}")
        else:
            # Adding a new device
            new_data["Status"] = "Added"
            df_new = pd.DataFrame([new_data])
            df_final = pd.concat([df_old, df_new], ignore_index=True)
            print(f"➕ New device added to registry.")
    else:
        # Creating the first log entry
        new_data["Status"] = "Added"
        df_final = pd.DataFrame([new_data])
        print(f"📄 Master inventory file created.")
            
    df_final.to_excel(filename, index=False)
    apply_formatting(filename)

def apply_formatting(filename):
    wb = load_workbook(filename)
    ws = wb.active
    # Professional blue header styling
    header_fill = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
    header_font = Font(bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 25
        
    wb.save(filename)
    print(f"✅ Success! Report saved and formatted: {filename}")

if __name__ == "__main__":
    data = collect_inventory()
    save_report(data)